"""
Migración masiva de especies, ejemplares y fotos al sistema Cactario Casa Molle.

Lee un Excel (hojas 'especies' y 'ejemplares') y una carpeta de fotos organizada
por slug, y crea todo vía la API del backend en el orden correcto:

    1) especie   -> POST /species/staff           (devuelve species_id)
    2) fotos esp -> POST /photos/especie/{id}      (multipart)
    3) ejemplar  -> POST /ejemplar/staff           (devuelve ejemplar_id)
    4) fotos ej  -> POST /photos/ejemplar/{id}     (multipart)

Autenticación: master key (--master-key) o OTP interactivo por email.

Ejemplos:
    python migrate.py --excel datos.xlsx --fotos ./fotos --api http://localhost:8000 --dry-run
    python migrate.py --excel datos.xlsx --fotos ./fotos --api http://localhost:8000 \
        --email staff@hotel.cl --master-key "LA_CLAVE"
"""
import argparse
import mimetypes
import re
import sys
from pathlib import Path

import requests
from openpyxl import load_workbook

# Campos de especie que se copian desde el Excel (el 'slug' se genera aparte).
ESPECIE_FIELDS = {
    "nombre_común", "scientific_name", "nombres_comunes", "habitat",
    "distribución", "estado_conservación", "categoría_de_conservación",
    "tipo_planta", "tipo_morfología", "floración", "cuidado", "usos",
    "historia_nombre", "historia_y_leyendas", "Endémica", "expectativa_vida",
}
EJEMPLAR_FIELDS = {
    "tamaño", "health_status", "nursery", "invoice_number", "purchase_date",
    "purchase_price", "sale_price", "age_months", "location",
}
NUMERIC_FIELDS = {"purchase_price", "sale_price", "age_months"}
BOOL_FIELDS = {"Endémica"}
TRUTHY = {"sí", "si", "true", "verdadero", "1", "x", "yes", "y"}
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp"}


def slugify(scientific_name):
    """
    Genera el slug igual que el WMS:
        scientific_name.toLowerCase().replace(/\\s+/g,'-').replace(/[^\\w\\-]/g,'')
    'Echinopsis chiloensis' -> 'echinopsis-chiloensis'
    """
    s = (scientific_name or "").strip().lower()
    s = re.sub(r"\s+", "-", s)
    s = re.sub(r"[^\w\-]", "", s, flags=re.ASCII)
    return s


class ApiClient:
    def __init__(self, base_url, dry_run=False):
        self.base = base_url.rstrip("/")
        self.dry_run = dry_run
        self.session = requests.Session()
        self.token = None

    # ---------- auth ----------
    def login_master_key(self, email, master_key):
        r = self.session.post(
            f"{self.base}/auth/master-key-login",
            json={"email": email, "master_key": master_key}, timeout=30,
        )
        r.raise_for_status()
        self._set_token(r.json()["access_token"])

    def login_otp(self, email):
        r = self.session.post(
            f"{self.base}/auth/request-otp", json={"email": email}, timeout=30
        )
        if r.status_code not in (200, 204):
            raise RuntimeError(f"request-otp falló ({r.status_code}): {r.text}")
        code = input(f"Código OTP enviado a {email}. Ingrésalo: ").strip()
        r = self.session.post(
            f"{self.base}/auth/verify-otp", json={"email": email, "code": code}, timeout=30
        )
        r.raise_for_status()
        self._set_token(r.json()["access_token"])

    def _set_token(self, token):
        self.token = token
        self.session.headers["Authorization"] = f"Bearer {token}"
        print("  Autenticado correctamente.")

    # ---------- recursos ----------
    def get_sectors_map(self):
        r = self.session.get(f"{self.base}/sectors/staff", timeout=30)
        r.raise_for_status()
        data = r.json()
        rows = data.get("data", data) if isinstance(data, dict) else data
        return {str(s["name"]).strip().lower(): s["id"] for s in rows if s.get("name")}

    def create_species(self, payload):
        if self.dry_run:
            print(f"  [dry-run] crearía especie slug={payload.get('slug')}")
            return {"id": -1}
        r = self.session.post(f"{self.base}/species/staff", json=payload, timeout=30)
        if r.status_code == 400 and "slug ya existe" in r.text:
            raise SlugExists(payload["slug"])
        r.raise_for_status()
        return r.json()

    def create_ejemplar(self, payload):
        if self.dry_run:
            print(f"  [dry-run] crearía ejemplar species_id={payload.get('species_id')}")
            return {"id": -1}
        r = self.session.post(f"{self.base}/ejemplar/staff", json=payload, timeout=30)
        r.raise_for_status()
        return r.json()

    def upload_photos(self, entity_type, entity_id, photo_paths):
        if not photo_paths:
            return 0
        if self.dry_run:
            print(f"  [dry-run] subiría {len(photo_paths)} foto(s) a {entity_type}/{entity_id}")
            return len(photo_paths)
        files = []
        opened = []
        try:
            for p in photo_paths:
                fh = open(p, "rb")
                opened.append(fh)
                mime = mimetypes.guess_type(p.name)[0] or "image/jpeg"
                files.append(("files", (p.name, fh, mime)))
            r = self.session.post(
                f"{self.base}/photos/{entity_type}/{entity_id}", files=files, timeout=120
            )
            r.raise_for_status()
            return r.json().get("count", len(photo_paths))
        finally:
            for fh in opened:
                fh.close()


class SlugExists(Exception):
    pass


def _clean(value, field):
    """Normaliza un valor de celda para el payload JSON."""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return None
    if field in BOOL_FIELDS:
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in TRUTHY
    if field in NUMERIC_FIELDS and value is not None:
        try:
            return float(value) if "." in str(value) else int(value)
        except (ValueError, TypeError):
            return None
    if field == "purchase_date" and value is not None:
        # openpyxl puede devolver datetime; tomamos solo la fecha ISO
        return str(value)[:10]
    return value


def read_sheet(wb, name, required=True):
    if name not in wb.sheetnames:
        if required:
            raise RuntimeError(f"El Excel no tiene la hoja '{name}'")
        return []
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for raw in rows[1:]:
        if all(c is None or str(c).strip() == "" for c in raw):
            continue  # fila vacía
        out.append(dict(zip(headers, raw)))
    return out


def sort_photos(folder):
    if not folder.is_dir():
        return []
    return sorted(
        [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in IMAGE_EXTS],
        key=lambda p: p.name.lower(),
    )


def main():
    ap = argparse.ArgumentParser(description="Migración masiva Cactario")
    ap.add_argument("--excel", required=True, help="Ruta al .xlsx con hojas especies/ejemplares")
    ap.add_argument("--fotos", help="Carpeta raíz de fotos organizada por slug")
    ap.add_argument("--api", default="http://localhost:8000", help="URL base del backend")
    ap.add_argument("--email", help="Email autorizado para login")
    ap.add_argument("--master-key", help="Clave maestra (login no interactivo)")
    ap.add_argument("--dry-run", action="store_true", help="Valida sin escribir nada")
    ap.add_argument("--skip-existing", action="store_true",
                    help="Si el slug ya existe, salta la especie en vez de abortar")
    args = ap.parse_args()

    fotos_root = Path(args.fotos) if args.fotos else None
    if fotos_root and not fotos_root.is_dir():
        sys.exit(f"Carpeta de fotos no encontrada: {fotos_root}")

    wb = load_workbook(args.excel, data_only=True)
    especies = read_sheet(wb, "especies")
    ejemplares = read_sheet(wb, "ejemplares", required=False)
    print(f"Leídas {len(especies)} especies y {len(ejemplares)} ejemplares del Excel.")

    # Agrupar ejemplares por especie, usando el slug derivado del nombre científico
    ej_por_slug = {}
    for ej in ejemplares:
        nombre = (ej.get("especie_cientifico") or "").strip()
        if nombre:
            ej_por_slug.setdefault(slugify(nombre), []).append(ej)

    # --- Autenticación ---
    client = ApiClient(args.api, dry_run=args.dry_run)
    if not args.dry_run:
        if not args.email:
            sys.exit("Falta --email para autenticarse.")
        print("Autenticando...")
        if args.master_key:
            client.login_master_key(args.email, args.master_key)
        else:
            client.login_otp(args.email)
        sectors_map = client.get_sectors_map()
        print(f"  {len(sectors_map)} sectores disponibles.")
    else:
        sectors_map = {}
        print("[dry-run] No se autentica ni se contacta la API.")

    stats = {"especies": 0, "ejemplares": 0, "fotos": 0, "saltadas": 0, "errores": 0}

    for esp in especies:
        scientific = (esp.get("scientific_name") or "").strip()
        if not scientific:
            print("ERROR: fila de especie sin 'scientific_name', se omite.")
            stats["errores"] += 1
            continue
        slug = slugify(scientific)  # mismo criterio que el WMS

        # Payload de especie (solo campos válidos)
        payload = {}
        for field in ESPECIE_FIELDS:
            if field in esp:
                payload[field] = _clean(esp[field], field)
        payload["slug"] = slug

        print(f"\n>> Especie: {scientific}  (slug: {slug})")
        try:
            created = client.create_species(payload)
        except SlugExists:
            if args.skip_existing:
                print("  slug ya existe -> SALTADA")
                stats["saltadas"] += 1
                continue
            sys.exit(f"  slug '{slug}' ya existe. Usa --skip-existing para omitir.")
        except requests.HTTPError as e:
            print(f"  ERROR creando especie: {e} :: {e.response.text if e.response else ''}")
            stats["errores"] += 1
            continue

        species_id = created["id"]
        stats["especies"] += 1

        # Fotos de la especie: fotos/<slug>/*.jpg
        if fotos_root:
            esp_photos = sort_photos(fotos_root / slug)
            stats["fotos"] += client.upload_photos("especie", species_id, esp_photos)

        # Ejemplares de esta especie
        for ej in ej_por_slug.get(slug, []):
            ej_payload = {"species_id": species_id}
            sector_nombre = (ej.get("sector_nombre") or "").strip()
            if sector_nombre:
                sid = sectors_map.get(sector_nombre.lower())
                if sid is None:
                    print(f"  AVISO: sector '{sector_nombre}' no existe; ejemplar en standby.")
                else:
                    ej_payload["sector_id"] = sid
            for field in EJEMPLAR_FIELDS:
                if field in ej:
                    ej_payload[field] = _clean(ej[field], field)

            try:
                ej_created = client.create_ejemplar(ej_payload)
            except requests.HTTPError as e:
                print(f"  ERROR creando ejemplar: {e} :: {e.response.text if e.response else ''}")
                stats["errores"] += 1
                continue
            stats["ejemplares"] += 1
            ejemplar_id = ej_created["id"]

            # Fotos del ejemplar: fotos/<slug>/ejemplares/<ref>/*.jpg
            ref = (ej.get("ref") or "").strip()
            if fotos_root and ref:
                ej_photos = sort_photos(fotos_root / slug / "ejemplares" / ref)
                stats["fotos"] += client.upload_photos("ejemplar", ejemplar_id, ej_photos)

    print("\n" + "=" * 50)
    print("RESUMEN")
    print(f"  Especies creadas : {stats['especies']}")
    print(f"  Ejemplares creados: {stats['ejemplares']}")
    print(f"  Fotos subidas    : {stats['fotos']}")
    print(f"  Especies saltadas: {stats['saltadas']}")
    print(f"  Errores          : {stats['errores']}")
    print("=" * 50)


if __name__ == "__main__":
    main()
