"""
Genera una plantilla Excel (plantilla_migracion.xlsx) lista para llenar con los
datos de especies y ejemplares que se van a migrar al sistema Cactario.

Uso:
    python generate_template.py

Crea un archivo con 3 hojas:
  - "especies"    : una fila por especie. El único campo obligatorio es 'scientific_name'
                    (el 'slug' se genera solo a partir de él, igual que en el WMS).
  - "ejemplares"  : una fila por ejemplar físico (ligado a la especie por 'especie_cientifico')
  - "instrucciones": cómo llenar el archivo y cómo organizar las fotos
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# --- Columnas de la hoja 'especies', en el MISMO orden que el formulario del WMS ---
# (ojo con tildes y la ñ en los nombres). NO incluye 'slug': se autogenera del nombre científico.
ESPECIE_COLS = [
    "scientific_name",            # OBLIGATORIO (Nombre Científico). De aquí sale el slug.
    "nombre_común",
    "estado_conservación",        # descripción libre
    "categoría_de_conservación",  # MENÚ: No amenazado | Preocupación menor | Protegido | En peligro de extinción
    "Endémica",                   # Sí / No
    "habitat",
    "cuidado",
    "nombres_comunes",            # separados por comas
    "tipo_planta",                # texto libre (ej: Cactácea)
    "tipo_morfología",            # MENÚ: Columnar | Redondo | Agave | Tallo plano | Otro
    "distribución",
    "expectativa_vida",
    "floración",
    "usos",
    "historia_nombre",
    "historia_y_leyendas",
]

# Hoja 'ejemplares'. 'especie_cientifico', 'ref' y 'sector_nombre' son auxiliares del script.
EJEMPLAR_COLS = [
    "especie_cientifico",  # AUX: debe coincidir con un 'scientific_name' de la hoja 'especies'
    "ref",                 # AUX: identificador local del ejemplar -> carpeta de sus fotos
    "sector_nombre",       # AUX: nombre del sector (se resuelve a sector_id). Vacío = standby
    "tamaño",              # XS | S | M | L | XL | XXL
    "health_status",
    "nursery",
    "invoice_number",
    "purchase_date",       # YYYY-MM-DD
    "purchase_price",
    "sale_price",
    "age_months",
    "location",
]

HEADER_FILL = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
AUX_FILL = PatternFill(start_color="F9A825", end_color="F9A825", fill_type="solid")
REQ_FILL = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
AUX_COLS = {"especie_cientifico", "ref", "sector_nombre"}
REQ_COLS = {"scientific_name"}


def _write_header(ws, cols):
    for idx, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=idx, value=col)
        cell.font = HEADER_FONT
        if col in REQ_COLS:
            cell.fill = REQ_FILL          # rojo = obligatorio
        elif col in AUX_COLS:
            cell.fill = AUX_FILL          # naranjo = auxiliar (no se guarda tal cual)
        else:
            cell.fill = HEADER_FILL       # verde = campo normal
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(14, len(col) + 4)
    ws.freeze_panes = "A2"


def main():
    wb = Workbook()

    ws_esp = wb.active
    ws_esp.title = "especies"
    _write_header(ws_esp, ESPECIE_COLS)
    # Fila de ejemplo (Quisco). Se puede borrar antes de migrar.
    ws_esp.append([
        "Echinopsis chiloensis", "Quisco",
        "Endémica de Chile, protegida por ley", "Preocupación menor", "Sí",
        "Laderas rocosas y quebradas secas de Chile central y norte chico",
        "Pleno sol, suelo con buen drenaje, riego escaso",
        "Quisco, Cardón, Cactus candelabro", "Cactácea", "Columnar",
        "Desde la Región de Atacama hasta el Maule", "Más de 100 años",
        "Primavera-verano; flores blancas nocturnas",
        "Ornamental; artesanía con su corteza seca",
        "Del mapudungun 'quisco'", "Planta emblemática del norte chico",
    ])

    ws_ej = wb.create_sheet("ejemplares")
    _write_header(ws_ej, EJEMPLAR_COLS)
    ws_ej.append([
        "Echinopsis chiloensis", "ej1", "Sector A", "L", "Sano",
        "Vivero Norte", "F-001", "2026-01-15", 12000, "", 24, "Fila 3",
    ])
    ws_ej.append([
        "Echinopsis chiloensis", "ej2", "Sector A", "M", "Sano",
        "Vivero Norte", "F-001", "2026-01-15", 9000, "", 18, "Fila 3",
    ])

    ws_help = wb.create_sheet("instrucciones")
    instrucciones = [
        ["INSTRUCCIONES DE MIGRACIÓN"],
        [""],
        ["LEYENDA DE COLORES EN LAS CABECERAS:"],
        ["   Rojo    = obligatorio (scientific_name)"],
        ["   Naranjo = auxiliar: enlaza datos pero NO se guarda tal cual (especie_cientifico, ref, sector_nombre)"],
        ["   Verde   = campo normal de la especie/ejemplar"],
        [""],
        ["1. Hoja 'especies': una fila por especie."],
        ["   - 'scientific_name' es lo ÚNICO obligatorio. El identificador (slug) se genera solo:"],
        ["     'Echinopsis chiloensis'  ->  'echinopsis-chiloensis'  (minúsculas, espacios por guiones)."],
        ["   - Respeta las tildes/ñ de los nombres de columna (nombre_común, distribución, etc.)."],
        ["   - 'categoría_de_conservación' (menú): No amenazado | Preocupación menor | Protegido | En peligro de extinción"],
        ["   - 'tipo_morfología' (menú): Columnar | Redondo | Agave | Tallo plano | Otro"],
        ["   - 'Endémica': escribe Sí o No."],
        ["   - 'tipo_planta' es texto libre (ej: Cactácea). Deja vacío lo que no tengas."],
        [""],
        ["2. Hoja 'ejemplares': una fila por ejemplar físico."],
        ["   - 'especie_cientifico' debe ser idéntico al 'scientific_name' de la hoja 'especies'."],
        ["   - 'ref' es un identificador libre tuyo (ej. ej1, ej2). Enlaza con la carpeta de fotos del ejemplar."],
        ["   - 'sector_nombre' debe existir ya como sector en el sistema. Vacío = ejemplar en standby."],
        ["   - 'tamaño' solo acepta: XS, S, M, L, XL, XXL."],
        ["   - 'purchase_date' en formato YYYY-MM-DD."],
        [""],
        ["3. FOTOS — organízalas por SLUG (el slug se deriva del nombre científico):"],
        [""],
        ["   fotos/"],
        ["     echinopsis-chiloensis/       <- slug de 'Echinopsis chiloensis' (fotos de la FICHA)"],
        ["       01_portada.jpg"],
        ["       02_floracion.jpg"],
        ["       ejemplares/"],
        ["         ej1/                      <- coincide con la columna 'ref'"],
        ["           01.jpg"],
        [""],
        ["   La PRIMERA foto en orden alfabético se marca como portada. Nómbralas 01_, 02_, ..."],
        [""],
        ["4. Ejecuta:  python migrate.py --excel datos.xlsx --fotos ./fotos --api http://localhost:8000"],
        ["   Usa --dry-run primero para validar sin escribir nada."],
    ]
    for row in instrucciones:
        ws_help.append(row)
    ws_help["A1"].font = Font(bold=True, size=14)
    ws_help.column_dimensions["A"].width = 100

    out = "plantilla_migracion.xlsx"
    wb.save(out)
    print(f"Plantilla creada: {out}")


if __name__ == "__main__":
    main()
